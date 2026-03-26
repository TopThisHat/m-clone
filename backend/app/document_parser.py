"""
Multi-format document text extraction.

Supported formats:
  - PDF (.pdf)          — pdfplumber (text + tables)
  - DOCX (.docx)        — python-docx (paragraphs + tables)
  - Excel (.xlsx, .xls) — openpyxl (all sheets, preserving table structure)
  - CSV/TSV (.csv, .tsv)— stdlib csv
  - Images (.png, .jpg, .jpeg, .gif, .webp) — GPT-4o vision
"""
from __future__ import annotations

import base64
import csv
import io
import logging

logger = logging.getLogger(__name__)

# Maximum file size accepted before extraction (100 MB)
_MAX_FILE_SIZE = 100 * 1024 * 1024
# Maximum pages processed per PDF
_MAX_PDF_PAGES = 500
# Maximum sheets processed per Excel workbook
_MAX_EXCEL_SHEETS = 50
# Maximum rows processed per sheet
_MAX_EXCEL_ROWS = 100_000
# Maximum image size for GPT-4o vision (20 MB)
_MAX_IMAGE_SIZE = 20 * 1024 * 1024


def extract_pdf(contents: bytes) -> str:
    """Extract text and tables from a PDF, preserving table structure."""
    if len(contents) > _MAX_FILE_SIZE:
        raise ValueError(f"PDF too large ({len(contents) // 1024 // 1024} MB). Maximum is {_MAX_FILE_SIZE // 1024 // 1024} MB.")

    import pdfplumber

    pages: list[str] = []
    with pdfplumber.open(io.BytesIO(contents)) as pdf:
        for i, page in enumerate(pdf.pages[:_MAX_PDF_PAGES]):
            parts: list[str] = []

            # Extract tables first so we can format them as markdown
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    if not table:
                        continue
                    md = _table_to_markdown(table)
                    if md.strip():
                        parts.append(md)

            # Extract remaining text
            text = page.extract_text()
            if text and text.strip():
                parts.append(text.strip())

            if parts:
                pages.append(f"[Page {i + 1}]\n" + "\n\n".join(parts))

    return "\n\n".join(pages)


def extract_docx(contents: bytes) -> str:
    """Extract text and tables from a DOCX document."""
    if len(contents) > _MAX_FILE_SIZE:
        raise ValueError(f"DOCX too large ({len(contents) // 1024 // 1024} MB). Maximum is {_MAX_FILE_SIZE // 1024 // 1024} MB.")

    import docx

    doc = docx.Document(io.BytesIO(contents))
    parts: list[str] = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

        if tag == "p":
            # Paragraph
            para = element
            text_parts = []
            for run in para.iter():
                if run.text:
                    text_parts.append(run.text)
            text = "".join(text_parts).strip()
            if text:
                parts.append(text)

        elif tag == "tbl":
            # Table — find the matching Table object
            for table in doc.tables:
                if table._tbl is element:
                    rows: list[list[str]] = []
                    for row in table.rows:
                        cells = [cell.text.strip() for cell in row.cells]
                        rows.append(cells)
                    md = _table_to_markdown(rows)
                    if md.strip():
                        parts.append(md)
                    break

    return "\n\n".join(parts)


def extract_excel(contents: bytes, filename: str = "") -> str:
    """Extract all sheets from an Excel workbook as markdown tables."""
    if len(contents) > _MAX_FILE_SIZE:
        raise ValueError(f"Excel file too large ({len(contents) // 1024 // 1024} MB). Maximum is {_MAX_FILE_SIZE // 1024 // 1024} MB.")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames[:_MAX_EXCEL_SHEETS]:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row_num, row in enumerate(ws.iter_rows(values_only=True)):
            if row_num >= _MAX_EXCEL_ROWS:
                break
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if rows:
            md = _table_to_markdown(rows)
            parts.append(f"## Sheet: {sheet_name}\n\n{md}")

    wb.close()
    return "\n\n".join(parts)


def extract_csv(contents: bytes, filename: str = "") -> str:
    """Extract a CSV/TSV file as a markdown table."""
    if len(contents) > _MAX_FILE_SIZE:
        raise ValueError(f"CSV too large ({len(contents) // 1024 // 1024} MB). Maximum is {_MAX_FILE_SIZE // 1024 // 1024} MB.")

    text = contents.decode("utf-8", errors="replace")

    # Detect delimiter
    delimiter = "\t" if filename.lower().endswith(".tsv") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)

    rows: list[list[str]] = []
    for row in reader:
        if any(cell.strip() for cell in row):
            rows.append(row)

    if not rows:
        return ""

    return _table_to_markdown(rows)


async def extract_image(contents: bytes, filename: str = "", mime_type: str = "") -> str:
    """Extract text from an image using GPT-4o vision."""
    if len(contents) > _MAX_IMAGE_SIZE:
        raise ValueError(f"Image too large ({len(contents) // 1024 // 1024} MB). Maximum is {_MAX_IMAGE_SIZE // 1024 // 1024} MB.")

    from app.openai_factory import get_openai_client

    if not mime_type:
        ext = filename.rsplit(".", 1)[-1].lower() if filename else "png"
        mime_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "gif": "image/gif", "webp": "image/webp"}
        mime_type = mime_map.get(ext, "image/png")

    b64 = base64.b64encode(contents).decode("utf-8")

    resp = await get_openai_client().chat.completions.create(
        model="gpt-4o",
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": (
                        "Extract ALL text, data, and information from this image. "
                        "If the image contains tables, lists of people, organizational charts, "
                        "or structured data, preserve the structure using markdown tables. "
                        "For each person or entity, include their name, title/role, and any "
                        "affiliations or relationships shown. Be thorough and precise."
                    ),
                },
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{mime_type};base64,{b64}"},
                },
            ],
        }],
        max_tokens=4000,
        timeout=60,
    )
    return resp.choices[0].message.content or ""


def _table_to_markdown(rows: list[list[str | None]]) -> str:
    """Convert a list of rows into a markdown table."""
    if not rows:
        return ""

    # Clean cells
    clean_rows = []
    for row in rows:
        clean_rows.append([str(cell).strip() if cell else "" for cell in row])

    # Normalize column count
    max_cols = max(len(r) for r in clean_rows)
    for row in clean_rows:
        while len(row) < max_cols:
            row.append("")

    # Build markdown
    lines = []
    # Header
    lines.append("| " + " | ".join(clean_rows[0]) + " |")
    lines.append("| " + " | ".join("---" for _ in clean_rows[0]) + " |")
    # Body
    for row in clean_rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


# ── Format detection ──────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {
    ".pdf", ".docx",
    ".xlsx", ".xls",
    ".csv", ".tsv",
    ".png", ".jpg", ".jpeg", ".gif", ".webp",
}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
TABLE_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}


def get_extension(filename: str) -> str:
    """Return lowercase extension including the dot."""
    if "." not in filename:
        return ""
    return "." + filename.rsplit(".", 1)[-1].lower()


def get_format_metadata(contents: bytes, filename: str, ext: str) -> dict:
    """Return lightweight metadata about a document without full extraction.

    Returns a dict with at least ``"type"`` and format-specific counts
    (pages, sheets, rows) where applicable.  Resources opened for
    inspection are always closed via ``finally`` blocks.
    """
    if ext == ".pdf":
        import pdfplumber

        pdf = pdfplumber.open(io.BytesIO(contents))
        try:
            return {"type": "pdf", "pages": len(pdf.pages)}
        finally:
            pdf.close()

    elif ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        try:
            return {"type": "xlsx", "sheets": len(wb.sheetnames)}
        finally:
            wb.close()

    elif ext in (".csv", ".tsv"):
        text = contents.decode("utf-8", errors="replace")
        delimiter = "\t" if ext == ".tsv" else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        row_count = sum(1 for row in reader if any(cell.strip() for cell in row))
        return {"type": "csv", "rows": row_count}

    elif ext == ".docx":
        return {"type": "docx"}

    elif ext in IMAGE_EXTENSIONS:
        return {"type": "image"}

    else:
        return {"type": "unknown"}


async def extract_text(contents: bytes, filename: str) -> str:
    """Route to the appropriate extractor based on file extension."""
    ext = get_extension(filename)

    if ext == ".pdf":
        return extract_pdf(contents)
    elif ext == ".docx":
        return extract_docx(contents)
    elif ext in (".xlsx", ".xls"):
        return extract_excel(contents, filename)
    elif ext in (".csv", ".tsv"):
        return extract_csv(contents, filename)
    elif ext in IMAGE_EXTENSIONS:
        return await extract_image(contents, filename)
    else:
        raise ValueError(f"Unsupported file format: {ext}")
